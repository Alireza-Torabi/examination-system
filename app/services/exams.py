import io
import json
import random
import re
import string
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models import Answer, Attempt, Choice, Exam, Question


def parse_questions_from_excel(file_stream):
    workbook = load_workbook(file_stream, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    normalized_headers = [h.strip().lower() if isinstance(h, str) else "" for h in headers]
    canonical_headers = []
    for h in normalized_headers:
        if h.startswith("question") and "image" in h:
            canonical_headers.append("question_image")
        elif h.startswith("question"):
            canonical_headers.append("question")
        elif h.startswith("type"):
            canonical_headers.append("type")
        elif h.startswith("option"):
            opt = h.replace(" ", "")
            match_img = re.match(r"option(\d+).*image", opt)
            match_txt = re.match(r"option(\d+)", opt)
            if match_img:
                canonical_headers.append(f"option{match_img.group(1)}_image")
            elif match_txt:
                canonical_headers.append(f"option{match_txt.group(1)}")
            else:
                canonical_headers.append("")
        elif h.startswith("correct"):
            canonical_headers.append("correct")
        elif h.startswith("reason"):
            canonical_headers.append("reason")
        else:
            canonical_headers.append(h)
    idx = {}
    for i, name in enumerate(canonical_headers):
        if name and name not in idx:
            idx[name] = i
    required_base = ["question", "type", "option1", "option2"]
    if any(r not in idx for r in required_base):
        raise ValueError(
            "Invalid template. Please download the provided template to prepare your Excel file."
        )
    has_correct = "correct" in idx
    has_reason = "reason" in idx

    option_text_headers = [h for h in canonical_headers if re.match(r"option\d+$", h)]
    option_image_headers = [h for h in canonical_headers if re.match(r"option\d+_image$", h)]
    option_text_headers.sort(key=lambda name: int(name.replace("option", "")) if name.replace("option", "").isdigit() else 99)
    option_image_headers.sort(key=lambda name: int(name.replace("option", "").replace("_image", "")) if name else 99)
    if len(option_text_headers) < 2:
        raise ValueError("At least two options are required (Option1, Option2).")
    if len(option_text_headers) > 6:
        raise ValueError("Only up to 6 options are supported.")

    questions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        question_text = row[idx["question"]]
        if not question_text:
            continue
        qtype = str(row[idx["type"]] or "").strip().lower()
        qtype = "multiple" if "multi" in qtype else "single"
        q_image_path = None
        if "question_image" in idx:
            q_image_path_raw = row[idx["question_image"]]
            q_image_path = str(q_image_path_raw).strip() if q_image_path_raw else None
        option_values = []
        option_numbers = []
        for opt_key in option_text_headers:
            val = row[idx[opt_key]]
            option_numbers.append(int(opt_key.replace("option", "")) if opt_key.replace("option", "").isdigit() else len(option_numbers) + 1)
            option_values.append(str(val).strip() if val else "")
        option_images = {}
        for opt_key in option_image_headers:
            num_str = opt_key.replace("option", "").replace("_image", "")
            num = int(num_str) if num_str.isdigit() else None
            if num is None:
                continue
            val = row[idx[opt_key]]
            option_images[num] = str(val).strip() if val else ""

        options = []
        for num, text_val in zip(option_numbers, option_values):
            img_val = option_images.get(num, "")
            if img_val and not text_val:
                raise ValueError(f"Question '{question_text}' has an image for option {num} but no text.")
            options.append({"text": text_val, "image_path": img_val or None})
        while options and options[-1]["text"] == "" and not options[-1]["image_path"]:
            options.pop()
        if len(options) < 2:
            raise ValueError(f"Question '{question_text}' must have at least two options.")
        if any(not opt["text"] for opt in options):
            raise ValueError(f"Question '{question_text}' has empty option gaps. Please fill options without gaps.")
        correct_indices = []
        if has_correct:
            raw = row[idx["correct"]] if "correct" in idx else None
            if raw:
                letters = str(raw).replace(" ", "").upper().split(",")
                letter_map = list(string.ascii_uppercase)
                for part in letters:
                    if part in letter_map and letter_map.index(part) < len(options):
                        correct_indices.append(letter_map.index(part))
        reason_val = ""
        if has_reason:
            reason_val = row[idx["reason"]] if "reason" in idx else ""
        questions.append(
            {
                "text": str(question_text),
                "qtype": qtype,
                "options": options,
                "correct": correct_indices,
                "reason": str(reason_val) if reason_val else None,
                "image_path": q_image_path,
            }
        )
    if not questions:
        raise ValueError("No questions found in the uploaded Excel file.")
    return questions


def create_questions(exam: Exam, question_defs: list[dict]):
    for q_def in question_defs:
        q = Question(
            exam=exam,
            text=q_def["text"],
            qtype=q_def["qtype"],
            tenant_id=exam.tenant_id,
            image_path=q_def.get("image_path"),
            reason=q_def.get("reason"),
            reason_image_path=q_def.get("reason_image_path"),
        )
        db.session.add(q)
        db.session.flush()
        correct = set(q_def.get("correct", []))
        for idx, opt in enumerate(q_def["options"]):
            opt_text = opt.get("text") if isinstance(opt, dict) else opt
            opt_image = opt.get("image_path") if isinstance(opt, dict) else None
            choice = Choice(
                question=q,
                text=opt_text,
                image_path=opt_image,
                is_correct=idx in correct,
                tenant_id=exam.tenant_id,
            )
            db.session.add(choice)


def attempt_end_time(attempt: Attempt):
    return attempt.started_at + timedelta(minutes=attempt.exam.duration_minutes)


def ensure_time_left(attempt: Attempt) -> bool:
    return datetime.utcnow() <= attempt_end_time(attempt)


def grade_attempt(attempt: Attempt):
    order = attempt.as_order_list()
    correct_count = 0
    for qid in order:
        question = db.session.get(Question, qid)
        if not question or question.tenant_id != attempt.tenant_id:
            continue
        correct_choices = {c.id for c in question.choices if c.is_correct}
        given = {
            ans.choice_id
            for ans in Answer.query.filter_by(attempt_id=attempt.id, question_id=question.id).all()
        }
        if given and given == correct_choices:
            correct_count += 1
    total = len(order) if order else 0
    percent = round((correct_count / total) * 100, 2) if total else 0
    attempt.num_correct = correct_count
    attempt.num_questions = total
    attempt.score_percent = percent
    attempt.submitted_at = datetime.utcnow()
    db.session.commit()


def export_exam_to_workbook(exam: Exam) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    max_opts = 6
    option_headers = []
    for i in range(1, max_opts + 1):
        option_headers.extend([f"Option{i}", f"Option{i}Image"])
    ws.append(
        [
            "Question",
            "QuestionImage",
            "Type (single/multiple)",
            *option_headers,
            "Correct (letters)",
            "Reason (optional)",
        ]
    )
    letter_map = list(string.ascii_uppercase)
    for q in exam.questions:
        options = list(q.choices)
        correct_letters = [
            letter_map[idx] for idx, c in enumerate(options) if c.is_correct and idx < len(letter_map)
        ]
        row = [q.text, q.image_path or "", q.qtype]
        for i in range(max_opts):
            if i < len(options):
                row.append(options[i].text)
                row.append(options[i].image_path or "")
            else:
                row.append("")
                row.append("")
        row.append(",".join(correct_letters))
        row.append(q.reason or "")
        ws.append(row)
    return wb
