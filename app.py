import io
import json
import os
import random
import re
import string
import sys
from datetime import datetime, timedelta, timezone
from functools import wraps
from zoneinfo import ZoneInfo

from flask import (
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from sqlalchemy import inspect, text
from openpyxl import Workbook, load_workbook
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from app import create_app
from app.config import Config
from app.extensions import db


TIMEZONE_OPTIONS = Config.TIMEZONE_OPTIONS

app = create_app()


class Tenant(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    slug = db.Column(db.String(80), unique=True, nullable=False)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # instructor or student
    full_name = db.Column(db.String(120))
    tenant_id = db.Column(db.Integer, db.ForeignKey("tenant.id"), nullable=False)
    tenant = db.relationship("Tenant")
    instructor_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    instructor = db.relationship("User", remote_side=[id], backref="students", foreign_keys=[instructor_id])
    timezone = db.Column(db.String(64))

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)


class Exam(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    start_at = db.Column(db.DateTime, nullable=False)
    end_at = db.Column(db.DateTime, nullable=False)
    duration_minutes = db.Column(db.Integer, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    tenant_id = db.Column(db.Integer, db.ForeignKey("tenant.id"), nullable=False)
    tenant = db.relationship("Tenant")
    creator = db.relationship("User", foreign_keys=[created_by])
    timezone = db.Column(db.String(64), nullable=False, default="UTC")
    deleted_at = db.Column(db.DateTime)
    question_limit = db.Column(db.Integer)
    is_closed = db.Column(db.Boolean, default=False)
    closed_at = db.Column(db.DateTime)

    questions = db.relationship("Question", backref="exam", cascade="all, delete-orphan")

    def is_active(self, now: datetime) -> bool:
        start = self.start_at if self.start_at.tzinfo else self.start_at.replace(tzinfo=timezone.utc)
        end = self.end_at if self.end_at.tzinfo else self.end_at.replace(tzinfo=timezone.utc)
        current = now if now.tzinfo else now.replace(tzinfo=timezone.utc)
        return (not self.is_closed) and start <= current <= end

    def has_answer_key(self) -> bool:
        return bool(self.questions) and all(any(choice.is_correct for choice in q.choices) for q in self.questions)


class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, db.ForeignKey("exam.id"), nullable=False)
    text = db.Column(db.Text, nullable=False)
    qtype = db.Column(db.String(20), nullable=False)  # single or multiple
    tenant_id = db.Column(db.Integer, db.ForeignKey("tenant.id"), nullable=False)
    tenant = db.relationship("Tenant")
    image_path = db.Column(db.String(300))
    reason = db.Column(db.Text)

    choices = db.relationship("Choice", backref="question", cascade="all, delete-orphan")


class Choice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    question_id = db.Column(db.Integer, db.ForeignKey("question.id"), nullable=False)
    text = db.Column(db.String(400), nullable=False)
    is_correct = db.Column(db.Boolean, default=False)
    tenant_id = db.Column(db.Integer, db.ForeignKey("tenant.id"), nullable=False)
    tenant = db.relationship("Tenant")


class Attempt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, db.ForeignKey("exam.id"), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    started_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    submitted_at = db.Column(db.DateTime)
    score_percent = db.Column(db.Float)
    num_correct = db.Column(db.Integer)
    num_questions = db.Column(db.Integer)
    question_order = db.Column(db.Text, nullable=False)  # JSON list of question IDs
    tenant_id = db.Column(db.Integer, db.ForeignKey("tenant.id"), nullable=False)
    tenant = db.relationship("Tenant")
    student = db.relationship("User")

    exam = db.relationship("Exam", backref="attempts")

    def as_order_list(self) -> list[int]:
        try:
            return json.loads(self.question_order)
        except json.JSONDecodeError:
            return []


class Answer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    attempt_id = db.Column(db.Integer, db.ForeignKey("attempt.id"), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey("question.id"), nullable=False)
    choice_id = db.Column(db.Integer, db.ForeignKey("choice.id"), nullable=False)
    tenant_id = db.Column(db.Integer, db.ForeignKey("tenant.id"), nullable=False)

    attempt = db.relationship("Attempt", backref="answers")
    choice = db.relationship("Choice")
    tenant = db.relationship("Tenant")


class ExamProgress(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, db.ForeignKey("exam.id"), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    tenant_id = db.Column(db.Integer, db.ForeignKey("tenant.id"), nullable=False)
    asked_questions = db.Column(db.Text, default="[]")  # JSON list of question IDs already asked in cycle

    exam = db.relationship("Exam")
    student = db.relationship("User")
    tenant = db.relationship("Tenant")

class ExamDeletionLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, nullable=False)
    exam_title = db.Column(db.String(200))
    instructor_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    tenant_id = db.Column(db.Integer, db.ForeignKey("tenant.id"), nullable=False)
    deleted_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    note = db.Column(db.Text)

    instructor = db.relationship("User", foreign_keys=[instructor_id])
    tenant = db.relationship("Tenant")


class AccessLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ip = db.Column(db.String(64))
    path = db.Column(db.String(400))
    method = db.Column(db.String(10))
    user_agent = db.Column(db.String(400))
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    tenant_id = db.Column(db.Integer, db.ForeignKey("tenant.id"))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User")
    tenant = db.relationship("Tenant")


def get_current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return db.session.get(User, uid)


@app.context_processor
def inject_user():
    return {"current_user": get_current_user(), "is_rtl": is_rtl_text}


def to_local(dt: datetime, tz_name: str) -> datetime:
    if not dt:
        return dt
    try:
        tz = ZoneInfo(tz_name)
    except Exception:  # pragma: no cover
        tz = timezone.utc
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(tz)


def local_to_utc(dt: datetime, tz_name: str) -> datetime:
    if not dt:
        return dt
    try:
        tz = ZoneInfo(tz_name)
    except Exception:  # pragma: no cover
        tz = timezone.utc
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=tz)
    else:
        dt = dt.astimezone(tz)
    return dt.astimezone(timezone.utc).replace(tzinfo=None)


def fmt_dt(dt: datetime) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def is_rtl_text(text: str) -> bool:
    if not text:
        return False
    return bool(re.search(r"[\u0600-\u06FF]", text))


def fmt_datetime_local_input(dt: datetime, tz_name: str) -> str:
    if not dt:
        return ""
    local = to_local(dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc), tz_name)
    return local.strftime("%Y-%m-%dT%H:%M")


def ensure_default_tenant():
    tenant = Tenant.query.filter_by(slug="default").first()
    if not tenant:
        tenant = Tenant(name="Default Tenant", slug="default")
        db.session.add(tenant)
        db.session.commit()
    return tenant


def ensure_column(table: str, column: str, ddl: str):
    inspector = inspect(db.engine)
    cols = [c["name"] for c in inspector.get_columns(table)]
    if column in cols:
        return False
    db.session.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {ddl}"))
    db.session.commit()
    return True


def migrate_schema():
    # Create missing tables first
    db.create_all()
    default_tenant = ensure_default_tenant()

    # Users
    if "user" in inspect(db.engine).get_table_names():
        added = ensure_column("user", "tenant_id", "INTEGER")
        if added:
            db.session.execute(text("UPDATE user SET tenant_id = :tid WHERE tenant_id IS NULL"), {"tid": default_tenant.id})
        ensure_column("user", "instructor_id", "INTEGER")
        ensure_column("user", "timezone", "VARCHAR(64)")

    # Exams
    if "exam" in inspect(db.engine).get_table_names():
        added = ensure_column("exam", "tenant_id", "INTEGER")
        if added:
            db.session.execute(text("UPDATE exam SET tenant_id = :tid WHERE tenant_id IS NULL"), {"tid": default_tenant.id})
        if ensure_column("exam", "timezone", "VARCHAR(64)"):
            db.session.execute(text("UPDATE exam SET timezone = 'UTC' WHERE timezone IS NULL"))
        ensure_column("exam", "deleted_at", "DATETIME")
        ensure_column("exam", "question_limit", "INTEGER")
        ensure_column("exam", "is_closed", "BOOLEAN")
        ensure_column("exam", "closed_at", "DATETIME")

    # Questions
    if "question" in inspect(db.engine).get_table_names():
        added = ensure_column("question", "tenant_id", "INTEGER")
        if added:
            db.session.execute(text("UPDATE question SET tenant_id = :tid WHERE tenant_id IS NULL"), {"tid": default_tenant.id})
        ensure_column("question", "image_path", "VARCHAR(300)")
        ensure_column("question", "reason", "TEXT")

    # Choices
    if "choice" in inspect(db.engine).get_table_names():
        added = ensure_column("choice", "tenant_id", "INTEGER")
        if added:
            db.session.execute(text("UPDATE choice SET tenant_id = :tid WHERE tenant_id IS NULL"), {"tid": default_tenant.id})

    # Attempts
    if "attempt" in inspect(db.engine).get_table_names():
        added = ensure_column("attempt", "tenant_id", "INTEGER")
        if added:
            db.session.execute(text("UPDATE attempt SET tenant_id = :tid WHERE tenant_id IS NULL"), {"tid": default_tenant.id})

    # Answers
    if "answer" in inspect(db.engine).get_table_names():
        added = ensure_column("answer", "tenant_id", "INTEGER")
        if added:
            db.session.execute(text("UPDATE answer SET tenant_id = :tid WHERE tenant_id IS NULL"), {"tid": default_tenant.id})

    # Access logs table (created via create_all)
    db.create_all()

    db.session.commit()


def login_required(role=None):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            user = get_current_user()
            if not user:
                flash("Please log in first.")
                return redirect(url_for("login"))
            if role:
                allowed = {role} if isinstance(role, str) else set(role)
                if user.role != "admin" and user.role not in allowed:
                    abort(403)
            return func(*args, **kwargs)

        return wrapper

    return decorator


def parse_datetime(value: str) -> datetime:
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def parse_questions_from_excel(file_stream):
    workbook = load_workbook(file_stream, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    normalized_headers = [h.strip().lower() if isinstance(h, str) else "" for h in headers]
    canonical_headers = []
    for h in normalized_headers:
        if h.startswith("question"):
            canonical_headers.append("question")
        elif h.startswith("type"):
            canonical_headers.append("type")
        elif h.startswith("option"):
            canonical_headers.append(h.split()[0])  # keep option1..optionN
        elif h.startswith("correct"):
            canonical_headers.append("correct")
        elif h.startswith("reason"):
            canonical_headers.append("reason")
        else:
            canonical_headers.append(h)
    # map first occurrence of each canonical header to its index
    idx = {}
    for i, name in enumerate(canonical_headers):
        if name and name not in idx:
            idx[name] = i
    required_base = ["question", "type", "option1", "option2"]
    if any(r not in idx for r in required_base):
        raise ValueError(
            "Invalid template. Please download the provided template to prepare your Excel file."
        )
    has_correct = "correct" in idx
    has_reason = "reason" in idx

    option_headers = [h for h in canonical_headers if h.startswith("option")]
    option_headers.sort(key=lambda name: int(name.replace("option", "")) if name.replace("option", "").isdigit() else 99)
    if len(option_headers) < 2:
        raise ValueError("At least two options are required (Option1, Option2).")

    questions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        question_text = row[idx["question"]]
        if not question_text:
            continue
        qtype = str(row[idx["type"]] or "").strip().lower()
        qtype = "multiple" if "multi" in qtype else "single"
        option_values = []
        for opt_key in option_headers:
            val = row[idx[opt_key]]
            option_values.append(str(val).strip() if val else "")
        while option_values and option_values[-1] == "":
            option_values.pop()
        if len(option_values) < 2:
            raise ValueError(f"Question '{question_text}' must have at least two options.")
        if "" in option_values:
            raise ValueError(f"Question '{question_text}' has empty option gaps. Please fill options without gaps.")
        options = option_values
        correct_indices = []
        if has_correct:
            raw = row[idx["correct"]] if "correct" in idx else None
            if raw:
                letters = str(raw).replace(" ", "").upper().split(",")
                letter_map = list(string.ascii_uppercase)
                for part in letters:
                    if part in letter_map and letter_map.index(part) < len(options):
                        correct_indices.append(letter_map.index(part))
        reason_val = ""
        if has_reason:
            reason_val = row[idx["reason"]] if "reason" in idx else ""
        questions.append(
            {
                "text": str(question_text),
                "qtype": qtype,
                "options": options,
                "correct": correct_indices,
                "reason": str(reason_val) if reason_val else None,
            }
        )
    if not questions:
        raise ValueError("No questions found in the uploaded Excel file.")
    return questions


def create_questions(exam: Exam, question_defs: list[dict]):
    for q_def in question_defs:
        q = Question(
            exam=exam,
            text=q_def["text"],
            qtype=q_def["qtype"],
            tenant_id=exam.tenant_id,
            image_path=q_def.get("image_path"),
            reason=q_def.get("reason"),
        )
        db.session.add(q)
        db.session.flush()
        correct = set(q_def.get("correct", []))
        for idx, text in enumerate(q_def["options"]):
            choice = Choice(question=q, text=text, is_correct=idx in correct, tenant_id=exam.tenant_id)
            db.session.add(choice)


def attempt_end_time(attempt: Attempt) -> datetime:
    return attempt.started_at + timedelta(minutes=attempt.exam.duration_minutes)


def ensure_time_left(attempt: Attempt) -> bool:
    return datetime.utcnow() <= attempt_end_time(attempt)


@app.route("/")
def index():
    user = get_current_user()
    if user:
        if user.role == "admin":
            return redirect(url_for("admin_dashboard"))
        return redirect(url_for("instructor_dashboard" if user.role == "instructor" else "student_dashboard"))
    return redirect(url_for("login"))


@app.before_request
def log_access():
    # Skip logging static files
    if request.path.startswith("/static"):
        return
    user = get_current_user()
    tenant_id = user.tenant_id if user else None
    ip = request.headers.get("X-Forwarded-For", request.remote_addr)
    try:
        entry = AccessLog(
            ip=ip,
            path=request.path[:400],
            method=request.method,
            user_agent=(request.headers.get("User-Agent") or "")[:400],
            user_id=user.id if user else None,
            tenant_id=tenant_id,
        )
        db.session.add(entry)
        db.session.commit()
    except Exception:
        db.session.rollback()


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session["user_id"] = user.id
            flash("Welcome back.")
            return redirect(url_for("index"))
        flash("Invalid credentials.")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("user_id", None)
    flash("Logged out.")
    return redirect(url_for("login"))


@app.route("/instructor")
@login_required(role="instructor")
def instructor_dashboard():
    user = get_current_user()
    if user.role == "admin":
        return redirect(url_for("admin_dashboard"))
    user_tz = user.timezone or "UTC"
    now_utc = datetime.now(timezone.utc)
    now_local = to_local(now_utc, user_tz)
    exams_raw = (
        Exam.query.filter_by(tenant_id=user.tenant_id, created_by=user.id)
        .filter(Exam.deleted_at.is_(None))
        .order_by(Exam.start_at.desc())
        .all()
    )
    exams = []
    for ex in exams_raw:
        start_local = fmt_dt(to_local(ex.start_at if ex.start_at.tzinfo else ex.start_at.replace(tzinfo=timezone.utc), user_tz))
        end_local = fmt_dt(to_local(ex.end_at if ex.end_at.tzinfo else ex.end_at.replace(tzinfo=timezone.utc), user_tz))
        exams.append({"obj": ex, "start_local": start_local, "end_local": end_local})
    return render_template("instructor_dashboard.html", exams=exams, now=now_local, user_timezone=user_tz)


@app.route("/admin")
@login_required(role="admin")
def admin_dashboard():
    tenant_filter = request.args.get("tenant_id", type=int)
    exam_filter = request.args.get("exam_id", type=int)
    tenants = Tenant.query.order_by(Tenant.name).all()
    admin_user = get_current_user()
    admin_tz = admin_user.timezone or "UTC"

    stats = {
        "tenants": Tenant.query.count(),
        "users": User.query.count(),
        "exams": Exam.query.filter(Exam.deleted_at.is_(None)).count(),
        "attempts": Attempt.query.count(),
    }

    users = User.query.order_by(User.tenant_id, User.role, User.username).all()

    exams_query = Exam.query.order_by(Exam.start_at.desc())
    if tenant_filter:
        exams_query = exams_query.filter_by(tenant_id=tenant_filter)
    exams = exams_query.filter(Exam.deleted_at.is_(None)).all()

    exam_options = Exam.query.order_by(Exam.title).all()

    return render_template(
        "admin_dashboard.html",
        stats=stats,
        tenants=tenants,
        users=users,
        exams=exams,
        tenant_filter=tenant_filter,
        exam_filter=exam_filter,
        exam_options=exam_options,
        admin_tz=admin_tz,
    )


@app.route("/logs")
@login_required(role="admin")
def admin_logs():
    admin_user = get_current_user()
    admin_tz = admin_user.timezone or "UTC"
    view = request.args.get("view", "app")
    access_logs = []
    deletion_logs = []
    attempt_logs = []

    if view == "access":
        access_logs = (
            AccessLog.query.order_by(AccessLog.created_at.desc())
            .limit(200)
            .all()
        )
    else:
        deletion_raw = ExamDeletionLog.query.order_by(ExamDeletionLog.deleted_at.desc()).limit(200).all()
        deletion_logs = [
            {"obj": log, "deleted_local": fmt_dt(to_local(log.deleted_at, admin_tz))}
            for log in deletion_raw
        ]
        attempt_raw = Attempt.query.order_by(Attempt.submitted_at.desc().nullslast()).limit(200).all()
        for att in attempt_raw:
            attempt_logs.append(
                {
                    "obj": att,
                    "started_local": fmt_dt(to_local(att.started_at, admin_tz)),
                    "submitted_local": fmt_dt(to_local(att.submitted_at, admin_tz)) if att.submitted_at else None,
                }
            )

    return render_template(
        "logs.html",
        access_logs=access_logs,
        deletion_logs=deletion_logs,
        attempt_logs=attempt_logs,
        view=view,
    )


def get_tenants_for_forms():
    return Tenant.query.order_by(Tenant.name).all()


@app.route("/admin/users/new", methods=["GET", "POST"])
@login_required(role="admin")
def admin_user_new():
    tenants = get_tenants_for_forms()
    instructors = User.query.filter_by(role="instructor").order_by(User.username).all()
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        full_name = request.form.get("full_name", "").strip()
        role = request.form.get("role", "student")
        password = request.form.get("password", "")
        tenant_id = request.form.get("tenant_id", type=int)
        instructor_id = request.form.get("instructor_id", type=int)
        timezone = request.form.get("timezone", "").strip() or None
        if not username or not password or not tenant_id:
            flash("Username, password, and tenant are required.")
            return redirect(request.url)
        if role not in {"admin", "instructor", "student"}:
            role = "student"
        if User.query.filter_by(username=username).first():
            flash("Username already exists.")
            return redirect(request.url)
        if role == "student" and instructor_id:
            instructor = db.session.get(User, instructor_id)
            if not instructor or instructor.role != "instructor" or instructor.tenant_id != tenant_id:
                flash("Instructor must be a valid instructor in the same tenant.")
                return redirect(request.url)
        db.session.add(
            User(
                username=username,
                full_name=full_name,
                role=role,
                password_hash=generate_password_hash(password),
                tenant_id=tenant_id,
                instructor_id=instructor_id if role == "student" else None,
                timezone=timezone,
            )
        )
        db.session.commit()
        flash("User created.")
        return redirect(url_for("admin_dashboard"))
    return render_template("user_form.html", tenants=tenants, user=None, instructors=instructors, timezone_options=TIMEZONE_OPTIONS)


@app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required(role="admin")
def admin_user_edit(user_id):
    user_obj = db.session.get(User, user_id)
    if not user_obj:
        abort(404)
    tenants = get_tenants_for_forms()
    instructors = User.query.filter_by(role="instructor").order_by(User.username).all()
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        if not username:
            flash("Username is required.")
            return redirect(request.url)
        existing = User.query.filter(User.username == username, User.id != user_obj.id).first()
        if existing:
            flash("Username already exists.")
            return redirect(request.url)
        user_obj.username = username
        user_obj.full_name = request.form.get("full_name", "").strip()
        user_obj.role = request.form.get("role", user_obj.role)
        tenant_id = request.form.get("tenant_id", type=int) or user_obj.tenant_id
        user_obj.tenant_id = tenant_id
        instructor_id = request.form.get("instructor_id", type=int)
        user_obj.timezone = request.form.get("timezone", "").strip() or None
        if user_obj.role == "student" and instructor_id:
            instructor = db.session.get(User, instructor_id)
            if not instructor or instructor.role != "instructor" or instructor.tenant_id != tenant_id:
                flash("Instructor must be a valid instructor in the same tenant.")
                return redirect(request.url)
            user_obj.instructor_id = instructor_id
        else:
            user_obj.instructor_id = None
        new_password = request.form.get("password", "")
        confirm_pw = request.form.get("password_confirm", "")
        if new_password:
            if confirm_pw and new_password != confirm_pw:
                flash("Passwords do not match.")
                return redirect(request.url)
            user_obj.password_hash = generate_password_hash(new_password)
        db.session.commit()
        flash("User updated.")
        return redirect(url_for("admin_dashboard"))
    return render_template("user_form.html", tenants=tenants, user=user_obj, instructors=instructors, timezone_options=TIMEZONE_OPTIONS)


@app.route("/admin/tenants/new", methods=["GET", "POST"])
@login_required(role="admin")
def admin_tenant_new():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        slug = request.form.get("slug", "").strip()
        if not name or not slug:
            flash("Name and slug are required.")
            return redirect(request.url)
        if Tenant.query.filter_by(slug=slug).first():
            flash("Slug already exists.")
            return redirect(request.url)
        db.session.add(Tenant(name=name, slug=slug))
        db.session.commit()
        flash("Tenant created.")
        return redirect(url_for("admin_dashboard"))
    return render_template("tenant_form.html")


@app.route("/excel-template")
@login_required(role="instructor")
def excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(
        [
            "Question",
            "Type (single/multiple)",
            "Option1",
            "Option2",
            "Option3",
            "Option4",
            "Option5",
            "Option6",
            "Correct (letters, e.g. A or A,C)",
            "Reason (optional)",
        ]
    )
    ws.append(["What is 2+2?", "single", "2", "3", "4", "5", "", "", "C", "Basic arithmetic."])
    ws.append(["Select prime numbers", "multiple", "2", "3", "4", "9", "11", "15", "A,E", "2 and 11 are prime."])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="exam_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/instructor/exams/new", methods=["GET", "POST"])
@login_required(role="instructor")
def create_exam():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "")
        start_at_raw = parse_datetime(request.form.get("start_at", ""))
        end_at_raw = parse_datetime(request.form.get("end_at", ""))
        duration_minutes = request.form.get("duration_minutes", "").strip()
        timezone = request.form.get("timezone", "UTC").strip() or "UTC"
        question_limit = request.form.get("question_limit", "").strip()
        upload = request.files.get("questions_file")

        if not title or not start_at_raw or not end_at_raw or not duration_minutes:
            flash("Title, times, and duration are required.")
            return redirect(request.url)
        try:
            duration_minutes = int(duration_minutes)
        except ValueError:
            flash("Duration must be a number of minutes.")
            return redirect(request.url)
        if question_limit:
            try:
                question_limit = int(question_limit)
                if question_limit <= 0:
                    raise ValueError()
            except ValueError:
                flash("Question count must be a positive number.")
                return redirect(request.url)
        else:
            question_limit = None
        start_at = local_to_utc(start_at_raw, timezone)
        end_at = local_to_utc(end_at_raw, timezone)
        if end_at <= start_at:
            flash("End time must be after start time.")
            return redirect(request.url)

        question_defs = []
        if upload and upload.filename:
            try:
                question_defs = parse_questions_from_excel(upload)
            except Exception as exc:  # pylint: disable=broad-except
                flash(str(exc))
                return redirect(request.url)

        exam = Exam(
            title=title,
            description=description,
            start_at=start_at,
            end_at=end_at,
            duration_minutes=duration_minutes,
            created_by=get_current_user().id,
            tenant_id=get_current_user().tenant_id,
            timezone=timezone,
            question_limit=question_limit,
        )
        db.session.add(exam)
        db.session.flush()
        if question_defs:
            create_questions(exam, question_defs)
        db.session.commit()
        if question_defs:
            flash("Exam created. Please set the correct answers.")
            return redirect(url_for("answer_key", exam_id=exam.id))
        flash("Exam created. Add questions from the UI.")
        return redirect(url_for("add_question", exam_id=exam.id))

    return render_template("exam_form.html", timezone_options=TIMEZONE_OPTIONS)


@app.route("/instructor/exams/<int:exam_id>/edit", methods=["GET", "POST"])
@login_required(role=["instructor", "admin"])
def edit_exam(exam_id):
    user = get_current_user()
    exam = db.session.get(Exam, exam_id)
    if not exam or exam.deleted_at:
        abort(404)
    if user.role != "admin" and (exam.created_by != user.id or exam.tenant_id != user.tenant_id):
        abort(403)
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "")
        start_at_raw = parse_datetime(request.form.get("start_at", ""))
        end_at_raw = parse_datetime(request.form.get("end_at", ""))
        duration_minutes = request.form.get("duration_minutes", "").strip()
        timezone = request.form.get("timezone", exam.timezone or "UTC").strip() or "UTC"
        question_limit = request.form.get("question_limit", "").strip()
        if not title or not start_at_raw or not end_at_raw or not duration_minutes:
            flash("Title, times, and duration are required.")
            return redirect(request.url)
        try:
            duration_minutes = int(duration_minutes)
        except ValueError:
            flash("Duration must be a number of minutes.")
            return redirect(request.url)
        if question_limit:
            try:
                question_limit = int(question_limit)
                if question_limit <= 0:
                    raise ValueError()
            except ValueError:
                flash("Question count must be a positive number.")
                return redirect(request.url)
        else:
            question_limit = None
        start_at = local_to_utc(start_at_raw, timezone)
        end_at = local_to_utc(end_at_raw, timezone)
        if end_at <= start_at:
            flash("End time must be after start time.")
            return redirect(request.url)

        exam.title = title
        exam.description = description
        exam.start_at = start_at
        exam.end_at = end_at
        exam.duration_minutes = duration_minutes
        exam.timezone = timezone
        exam.question_limit = question_limit
        db.session.commit()
        flash("Exam updated.")
        return redirect(url_for("instructor_dashboard"))

    start_val = fmt_datetime_local_input(exam.start_at, exam.timezone or "UTC")
    end_val = fmt_datetime_local_input(exam.end_at, exam.timezone or "UTC")
    return render_template(
        "exam_edit.html",
        exam=exam,
        start_val=start_val,
        end_val=end_val,
        timezone_options=TIMEZONE_OPTIONS,
    )


@app.route("/instructor/exams/<int:exam_id>/answers", methods=["GET", "POST"])
@login_required(role="instructor")
def answer_key(exam_id):
    exam = db.session.get(Exam, exam_id)
    user = get_current_user()
    if not exam:
        abort(404)
    if exam.deleted_at:
        flash("This exam was deleted.")
        return redirect(url_for("instructor_dashboard"))
    if user.role != "admin" and (exam.created_by != user.id or exam.tenant_id != user.tenant_id):
        abort(404)
    if request.method == "POST":
        for question in exam.questions:
            selected_ids = request.form.getlist(f"q_{question.id}")
            if question.qtype == "single" and len(selected_ids) > 1:
                selected_ids = selected_ids[:1]
            selected_ids_set = {int(sid) for sid in selected_ids}
            for choice in question.choices:
                choice.is_correct = choice.id in selected_ids_set
        db.session.commit()
        flash("Answer key saved.")
        return redirect(url_for("answer_key", exam_id=exam.id))
    return render_template("answer_key.html", exam=exam)


@app.route("/instructor/exams/<int:exam_id>/questions/new", methods=["GET", "POST"])
@login_required(role="instructor")
def add_question(exam_id):
    user = get_current_user()
    exam = db.session.get(Exam, exam_id)
    if not exam:
        abort(404)
    if exam.deleted_at:
        flash("This exam was deleted.")
        return redirect(url_for("instructor_dashboard"))
    if user.role != "admin" and (exam.created_by != user.id or exam.tenant_id != user.tenant_id):
        abort(404)
    if request.method == "POST":
        text = request.form.get("text", "").strip()
        qtype = request.form.get("qtype", "single")
        qtype = "multiple" if qtype == "multiple" else "single"
        image_file = request.files.get("image")
        reason = request.form.get("reason", "").strip()
        option_fields = []
        for idx in range(1, 7):
            val = request.form.get(f"option{idx}", "").strip()
            if val:
                option_fields.append((idx, val))
        if not text:
            flash("Question text is required.")
            return redirect(request.url)
        if len(option_fields) < 2:
            flash("At least two options are required.")
            return redirect(request.url)
        correct_raw = {int(v) for v in request.form.getlist("correct") if v.isdigit()}
        options = []
        correct_indices = set()
        for idx, (field_idx, val) in enumerate(option_fields):
            options.append(val)
            if field_idx in correct_raw:
                correct_indices.add(idx)
        if not correct_indices:
            flash("Please select at least one correct answer.")
            return redirect(request.url)
        if qtype == "single":
            first = sorted(correct_indices)[0]
            correct_indices = {first}
        image_path = None
        if image_file and image_file.filename:
            filename = secure_filename(image_file.filename)
            if not filename.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
                flash("Only image files are allowed (png, jpg, jpeg, gif, webp).")
                return redirect(request.url)
            upload_dir = current_app.config["UPLOAD_FOLDER"]
            os.makedirs(upload_dir, exist_ok=True)
            stored_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{filename}"
            file_path = os.path.join(upload_dir, stored_name)
            image_file.save(file_path)
            image_path = f"uploads/{stored_name}"
        question = Question(
            exam=exam,
            text=text,
            qtype=qtype,
            tenant_id=exam.tenant_id,
            image_path=image_path,
            reason=reason or None,
        )
        db.session.add(question)
        db.session.flush()
        for idx, opt_text in enumerate(options):
            db.session.add(
                Choice(
                    question=question,
                    text=opt_text,
                    is_correct=idx in correct_indices,
                    tenant_id=exam.tenant_id,
                )
            )
        db.session.commit()
        flash("Question added.")
        action = request.form.get("action", "add_more")
        if action == "finish":
            return redirect(url_for("answer_key", exam_id=exam.id))
        return redirect(request.url)
    return render_template("question_form.html", exam=exam)


@app.route("/instructor/questions/<int:question_id>/edit", methods=["GET", "POST"])
@login_required(role=["instructor", "admin"])
def edit_question(question_id):
    user = get_current_user()
    question = db.session.get(Question, question_id)
    if not question or question.exam.deleted_at:
        abort(404)
    exam = question.exam
    if user.role != "admin" and (exam.created_by != user.id or exam.tenant_id != user.tenant_id):
        abort(404)
    if request.method == "POST":
        text = request.form.get("text", "").strip()
        qtype = request.form.get("qtype", "single")
        qtype = "multiple" if qtype == "multiple" else "single"
        image_file = request.files.get("image")
        remove_image = request.form.get("remove_image") == "on"
        reason = request.form.get("reason", "").strip()
        option_fields = []
        for idx in range(1, 7):
            val = request.form.get(f"option{idx}", "").strip()
            if val:
                option_fields.append((idx, val))
        if not text:
            flash("Question text is required.")
            return redirect(request.url)
        if len(option_fields) < 2:
            flash("At least two options are required.")
            return redirect(request.url)
        correct_raw = {int(v) for v in request.form.getlist("correct") if v.isdigit()}
        options = []
        correct_indices = set()
        for idx, (field_idx, val) in enumerate(option_fields):
            options.append(val)
            if field_idx in correct_raw:
                correct_indices.add(idx)
        if not correct_indices:
            flash("Please select at least one correct answer.")
            return redirect(request.url)
        if qtype == "single":
            first = sorted(correct_indices)[0]
            correct_indices = {first}

        image_path = question.image_path
        if remove_image:
            image_path = None
        if image_file and image_file.filename:
            filename = secure_filename(image_file.filename)
            if not filename.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
                flash("Only image files are allowed (png, jpg, jpeg, gif, webp).")
                return redirect(request.url)
            upload_dir = current_app.config["UPLOAD_FOLDER"]
            os.makedirs(upload_dir, exist_ok=True)
            stored_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{filename}"
            file_path = os.path.join(upload_dir, stored_name)
            image_file.save(file_path)
            image_path = f"uploads/{stored_name}"

        # Replace choices and answers
        Answer.query.filter_by(question_id=question.id).delete()
        Choice.query.filter_by(question_id=question.id).delete()
        db.session.flush()

        question.text = text
        question.qtype = qtype
        question.image_path = image_path
        question.reason = reason or None
        db.session.add(question)
        db.session.flush()
        for idx, opt_text in enumerate(options):
            db.session.add(
                Choice(
                    question=question,
                    text=opt_text,
                    is_correct=idx in correct_indices,
                    tenant_id=exam.tenant_id,
                )
            )
        db.session.commit()
        flash("Question updated.")
        return redirect(url_for("add_question", exam_id=exam.id))

    # Prepare form defaults
    options = question.choices
    letter_map = ["A", "B", "C", "D", "E", "F"]
    correct_indices = [idx for idx, c in enumerate(options) if c.is_correct]
    return render_template(
        "question_edit.html",
        question=question,
        exam=exam,
        options=options,
        correct_indices=correct_indices,
        letter_map=letter_map,
    )


@app.route("/settings", methods=["GET", "POST"])
@login_required()
def settings():
    user = get_current_user()
    if request.method == "POST":
        tz = request.form.get("timezone", "").strip() or None
        user.timezone = tz
        db.session.commit()
        flash("Settings saved.")
        return redirect(url_for("settings"))
    return render_template("settings.html", user=user, timezone_options=TIMEZONE_OPTIONS)


@app.route("/settings/password", methods=["GET", "POST"])
@login_required()
def change_password():
    user = get_current_user()
    if request.method == "POST":
        current_pw = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "")
        confirm_pw = request.form.get("confirm_password", "")
        if not new_pw:
            flash("New password is required.")
            return redirect(request.url)
        if not user.check_password(current_pw):
            flash("Current password is incorrect.")
            return redirect(request.url)
        if not confirm_pw or new_pw != confirm_pw:
            flash("Passwords do not match.")
            return redirect(request.url)
        user.password_hash = generate_password_hash(new_pw)
        db.session.commit()
        flash("Password updated.")
        return redirect(url_for("settings"))
    return render_template("change_password.html", show_current=True)


@app.route("/instructor/exams/<int:exam_id>/delete", methods=["POST"])
@login_required(role=["instructor", "admin"])
def delete_exam(exam_id):
    user = get_current_user()
    exam = db.session.get(Exam, exam_id)
    if not exam or exam.deleted_at:
        flash("Exam not found or already deleted.")
        return redirect(url_for("instructor_dashboard"))
    if user.role != "admin" and (exam.created_by != user.id or exam.tenant_id != user.tenant_id):
        abort(404)
    exam.deleted_at = datetime.utcnow()
    db.session.add(
        ExamDeletionLog(
            exam_id=exam.id,
            exam_title=exam.title,
            instructor_id=user.id,
            tenant_id=exam.tenant_id,
            note="Deleted by instructor",
        )
    )
    db.session.commit()
    flash("Exam deleted and logged for admin review.")
    return redirect(url_for("instructor_dashboard"))


@app.route("/instructor/exams/<int:exam_id>/toggle_close", methods=["POST"])
@login_required(role="instructor")
def toggle_close_exam(exam_id):
    user = get_current_user()
    exam = db.session.get(Exam, exam_id)
    if not exam:
        abort(404)
    if user.role != "admin" and (exam.created_by != user.id or exam.tenant_id != user.tenant_id):
        abort(404)
    if exam.deleted_at:
        flash("Exam was deleted.")
        return redirect(url_for("instructor_dashboard"))
    exam.is_closed = not exam.is_closed
    exam.closed_at = datetime.utcnow() if exam.is_closed else None
    db.session.commit()
    flash("Exam closed." if exam.is_closed else "Exam reopened.")
    return redirect(url_for("instructor_dashboard"))


@app.route("/student")
@login_required(role="student")
def student_dashboard():
    user = get_current_user()
    user_tz = user.timezone or "UTC"
    now_utc = datetime.now(timezone.utc)
    now_local = to_local(now_utc, user_tz)
    user = get_current_user()
    exams_query = Exam.query.filter_by(tenant_id=user.tenant_id).filter(Exam.deleted_at.is_(None))
    if user.instructor_id:
        exams_query = exams_query.filter_by(created_by=user.instructor_id)
    else:
        exams_query = exams_query.filter_by(created_by=None)  # no exams will match
    exams = exams_query.order_by(Exam.start_at.asc()).all()
    attempts = Attempt.query.filter_by(student_id=user.id).all()
    attempts_by_exam = {a.exam_id: a for a in attempts}
    exam_views = []
    for ex in exams:
        attempt = attempts_by_exam.get(ex.id)
        has_key = ex.has_answer_key()
        status = "blocked"
        can_start = False
        start_utc = ex.start_at if ex.start_at.tzinfo else ex.start_at.replace(tzinfo=timezone.utc)
        end_utc = ex.end_at if ex.end_at.tzinfo else ex.end_at.replace(tzinfo=timezone.utc)
        start_local = to_local(start_utc, user_tz)
        end_local = to_local(end_utc, user_tz)
        countdown_seconds = max(0, int((start_utc - now_utc).total_seconds()))
        if not has_key:
            status = "not_ready"
        elif now_utc < start_utc:
            status = "upcoming"
        elif ex.is_closed:
            status = "closed"
        else:
            # after start time, always allow retakes
            if attempt and attempt.submitted_at:
                status = "completed_active"
            elif attempt and not attempt.submitted_at:
                status = "active"
            else:
                status = "active"
            can_start = True
        exam_views.append(
            {
                "exam": ex,
                "attempt": attempt,
                "status": status,
                "can_start": can_start,
                "countdown_seconds": countdown_seconds,
                "start_local": fmt_dt(start_local),
                "end_local": fmt_dt(end_local),
            }
        )
    return render_template(
        "student_dashboard.html",
        exam_views=exam_views,
        now=now_local,
        user_timezone=user_tz,
        attempts=attempts,
    )


@app.route("/instructor/exams/<int:exam_id>/results")
@login_required(role=["instructor", "admin"])
def instructor_exam_results(exam_id):
    user = get_current_user()
    exam = db.session.get(Exam, exam_id)
    if not exam or exam.deleted_at:
        abort(404)
    if user.role != "admin" and (exam.created_by != user.id or exam.tenant_id != user.tenant_id):
        abort(403)
    user_tz = user.timezone or "UTC"
    attempts_raw = Attempt.query.filter_by(exam_id=exam.id).order_by(Attempt.started_at.desc()).all()
    attempts = []
    for att in attempts_raw:
        attempts.append(
            {
                "obj": att,
                "student": att.student,
                "started_local": fmt_dt(to_local(att.started_at, user_tz)),
                "submitted_local": fmt_dt(to_local(att.submitted_at, user_tz)),
            }
        )
    return render_template(
        "exam_results.html",
        exam=exam,
        attempts=attempts,
        user_timezone=user_tz,
    )


def export_exam_to_workbook(exam: Exam) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    max_opts = max((len(q.choices) for q in exam.questions), default=4)
    max_opts = max(max_opts, 4)
    option_headers = [f"Option{i}" for i in range(1, max_opts + 1)]
    ws.append(["Question", "Type (single/multiple)", *option_headers, "Correct (letters)", "Reason (optional)"])
    letter_map = list(string.ascii_uppercase)
    for q in exam.questions:
        options = [c.text for c in q.choices]
        correct_letters = [
            letter_map[idx] for idx, c in enumerate(q.choices) if c.is_correct and idx < len(letter_map)
        ]
        row = [q.text, q.qtype]
        for i in range(max_opts):
            row.append(options[i] if i < len(options) else "")
        row.append(",".join(correct_letters))
        row.append(q.reason or "")
        ws.append(row)
    return wb


@app.route("/instructor/exams/<int:exam_id>/export")
@login_required(role=["instructor", "admin"])
def export_exam(exam_id):
    user = get_current_user()
    exam = db.session.get(Exam, exam_id)
    if not exam or exam.deleted_at:
        abort(404)
    if user.role != "admin" and (exam.created_by != user.id or exam.tenant_id != user.tenant_id):
        abort(403)
    wb = export_exam_to_workbook(exam)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"exam_{exam.id}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/exam/<int:exam_id>/start")
@login_required(role="student")
def start_exam(exam_id):
    exam = db.session.get(Exam, exam_id)
    if not exam:
        abort(404)
    user = get_current_user()
    if user.role != "admin":
        if not user.instructor_id or exam.created_by != user.instructor_id:
            abort(403)
    if exam.deleted_at:
        flash("This exam was deleted.")
        return redirect(url_for("student_dashboard"))
    now = datetime.now(timezone.utc)
    start_utc = exam.start_at if exam.start_at.tzinfo else exam.start_at.replace(tzinfo=timezone.utc)
    if now < start_utc:
        flash("This exam has not started yet.")
        return redirect(url_for("student_dashboard"))
    if exam.is_closed:
        flash("This exam has been closed by the instructor.")
        return redirect(url_for("student_dashboard"))
    if exam.tenant_id != user.tenant_id:
        abort(403)
    if not exam.has_answer_key():
        flash("This exam is not yet ready. Please try later.")
        return redirect(url_for("student_dashboard"))

    # allow multiple attempts; resume only if an attempt is in progress
    attempt = Attempt.query.filter_by(exam_id=exam.id, student_id=user.id, submitted_at=None).first()
    if attempt is None:
        progress = ExamProgress.query.filter_by(exam_id=exam.id, student_id=user.id).first()
        if not progress:
            progress = ExamProgress(
                exam_id=exam.id,
                student_id=user.id,
                tenant_id=exam.tenant_id,
                asked_questions="[]",
            )
            db.session.add(progress)
            db.session.commit()
        try:
            asked_set = set(json.loads(progress.asked_questions or "[]"))
        except json.JSONDecodeError:
            asked_set = set()
        all_qids = [q.id for q in exam.questions]
        if not all_qids:
            flash("No questions available for this exam.")
            return redirect(url_for("student_dashboard"))
        # reset if cycle completed
        if len(asked_set) >= len(all_qids):
            asked_set = set()
        available = [qid for qid in all_qids if qid not in asked_set]
        if not available:
            asked_set = set()
            available = all_qids.copy()
        random.shuffle(available)
        if exam.question_limit and exam.question_limit > 0:
            selected = available[: exam.question_limit]
        else:
            selected = available
        if not selected:
            flash("No questions available to start the exam.")
            return redirect(url_for("student_dashboard"))
        # mark selected as asked in progress
        asked_set.update(selected)
        progress.asked_questions = json.dumps(list(asked_set))
        db.session.add(progress)

        attempt = Attempt(
            exam=exam,
            student_id=user.id,
            started_at=datetime.utcnow(),
            question_order=json.dumps(selected),
            num_questions=len(selected),
            tenant_id=exam.tenant_id,
        )
        db.session.add(attempt)
        db.session.commit()
    return redirect(url_for("show_question", attempt_id=attempt.id, index=1))


def get_attempt_or_404(attempt_id: int) -> Attempt:
    attempt = db.session.get(Attempt, attempt_id)
    if not attempt:
        abort(404)
    user = get_current_user()
    if attempt.student_id != user.id or attempt.tenant_id != user.tenant_id:
        abort(403)
    return attempt


@app.route("/attempt/<int:attempt_id>/question/<int:index>", methods=["GET", "POST"])
@login_required(role="student")
def show_question(attempt_id, index):
    attempt = get_attempt_or_404(attempt_id)
    if attempt.submitted_at:
        flash("Exam already submitted.")
        return redirect(url_for("view_result", attempt_id=attempt.id))

    order = attempt.as_order_list()
    if index < 1 or index > len(order):
        abort(404)
    question_id = order[index - 1]
    question = db.session.get(Question, question_id)
    if not question or question.exam_id != attempt.exam_id or question.tenant_id != attempt.tenant_id:
        abort(404)

    end_time = attempt_end_time(attempt)
    time_left_seconds = int((end_time - datetime.utcnow()).total_seconds())
    total_seconds = int((end_time - attempt.started_at).total_seconds())
    per_question_seconds = 0
    if len(order) > 0:
        per_question_seconds = max(1, total_seconds // len(order))
    if time_left_seconds <= 0:
        flash("Time is up. Auto-submitting your attempt with unanswered questions marked as empty.")
        grade_attempt(attempt)
        return redirect(url_for("view_result", attempt_id=attempt.id))

    existing_answers = Answer.query.filter_by(attempt_id=attempt.id, question_id=question.id).all()
    selected_ids = {ans.choice_id for ans in existing_answers}

    if request.method == "POST":
        if not ensure_time_left(attempt):
            flash("Time expired. Auto-submitting your attempt with unanswered questions marked as empty.")
            grade_attempt(attempt)
            return redirect(url_for("view_result", attempt_id=attempt.id))
        selected = request.form.getlist("choice")
        Answer.query.filter_by(attempt_id=attempt.id, question_id=question.id).delete()
        db.session.commit()
        for sid in selected:
            choice_obj = db.session.get(Choice, int(sid))
            if choice_obj and choice_obj.question_id == question.id:
                db.session.add(
                    Answer(
                        attempt=attempt,
                        question_id=question.id,
                        choice_id=choice_obj.id,
                        tenant_id=attempt.tenant_id,
                    )
                )
        db.session.commit()
        action = request.form.get("action", "next")
        if action == "previous" and index > 1:
            return redirect(url_for("show_question", attempt_id=attempt.id, index=index - 1))
        if action == "review":
            return redirect(url_for("review_attempt", attempt_id=attempt.id))
        next_index = index + 1
        if next_index > len(order):
            return redirect(url_for("review_attempt", attempt_id=attempt.id))
        return redirect(url_for("show_question", attempt_id=attempt.id, index=next_index))

    return render_template(
        "question.html",
        attempt=attempt,
        question=question,
        index=index,
        total=len(order),
        selected_ids=selected_ids,
        time_left_seconds=time_left_seconds,
        total_seconds=total_seconds,
        per_question_seconds=per_question_seconds,
    )


def grade_attempt(attempt: Attempt):
    order = attempt.as_order_list()
    correct_count = 0
    for qid in order:
        question = db.session.get(Question, qid)
        if not question or question.tenant_id != attempt.tenant_id:
            continue
        correct_choices = {c.id for c in question.choices if c.is_correct}
        given = {
            ans.choice_id
            for ans in Answer.query.filter_by(attempt_id=attempt.id, question_id=question.id).all()
        }
        if given and given == correct_choices:
            correct_count += 1
    total = len(order) if order else 0
    percent = round((correct_count / total) * 100, 2) if total else 0
    attempt.num_correct = correct_count
    attempt.num_questions = total
    attempt.score_percent = percent
    attempt.submitted_at = datetime.utcnow()
    db.session.commit()


@app.route("/attempt/<int:attempt_id>/review")
@login_required(role="student")
def review_attempt(attempt_id):
    attempt = get_attempt_or_404(attempt_id)
    order = attempt.as_order_list()
    questions = []
    for qid in order:
        q_obj = db.session.get(Question, qid)
        if q_obj and q_obj.tenant_id == attempt.tenant_id:
            questions.append(q_obj)
    answers_map = {}
    for ans in Answer.query.filter_by(attempt_id=attempt.id).all():
        answers_map.setdefault(ans.question_id, set()).add(ans.choice_id)
    return render_template(
        "review.html",
        attempt=attempt,
        questions=questions,
        answers_map=answers_map,
        time_left_seconds=max(0, int((attempt_end_time(attempt) - datetime.utcnow()).total_seconds())),
    )


@app.route("/attempt/<int:attempt_id>/submit", methods=["POST"])
@login_required(role="student")
def submit_attempt(attempt_id):
    attempt = get_attempt_or_404(attempt_id)
    if attempt.submitted_at:
        return redirect(url_for("view_result", attempt_id=attempt.id))
    grade_attempt(attempt)
    flash("Exam submitted.")
    return redirect(url_for("view_result", attempt_id=attempt.id))


@app.route("/attempt/<int:attempt_id>/result")
@login_required(role="student")
def view_result(attempt_id):
    attempt = get_attempt_or_404(attempt_id)
    if not attempt.submitted_at:
        flash("Please submit your exam first.")
        return redirect(url_for("review_attempt", attempt_id=attempt.id))
    order = attempt.as_order_list()
    questions = []
    for qid in order:
        q_obj = db.session.get(Question, qid)
        if q_obj and q_obj.tenant_id == attempt.tenant_id:
            questions.append(q_obj)
    answers_map = {}
    for ans in Answer.query.filter_by(attempt_id=attempt.id).all():
        answers_map.setdefault(ans.question_id, set()).add(ans.choice_id)
    return render_template("result.html", attempt=attempt, questions=questions, answers_map=answers_map)


def init_db():
    migrate_schema()
    tenant = ensure_default_tenant()
    if not User.query.filter_by(username="admin").first():
        admin_user = User(
            username="admin",
            full_name="Admin",
            role="admin",
            password_hash=generate_password_hash("admin123"),
            tenant_id=tenant.id,
        )
        db.session.add(admin_user)
    if not User.query.filter_by(username="instructor").first():
        instructor = User(
            username="instructor",
            full_name="Default Instructor",
            role="instructor",
            password_hash=generate_password_hash("instructor123"),
            tenant_id=tenant.id,
        )
        db.session.add(instructor)
    if not User.query.filter_by(username="student1").first():
        student = User(
            username="student1",
            full_name="Student One",
            role="student",
            password_hash=generate_password_hash("student123"),
            tenant_id=tenant.id,
            instructor_id=User.query.filter_by(username="instructor").first().id if User.query.filter_by(username="instructor").first() else None,
        )
        db.session.add(student)
    db.session.commit()
    print("Database initialized. Default logins: instructor/instructor123 and student1/student123.")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "initdb":
        with app.app_context():
            init_db()
    else:
        with app.app_context():
            migrate_schema()
        app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
